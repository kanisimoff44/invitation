"""Формирование xlsx-выгрузки списка гостей."""
from __future__ import annotations

import io
from datetime import timezone
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app import config
from app.models import RSVP

_HEADER = ["№", "Имя и фамилия", "Доп. гостей", "Всего человек", "Обновлено"]


def build_guests_xlsx(db: Session, *, attending_only: bool = True) -> bytes:
    """Собирает xlsx со списком гостей. По умолчанию — только те, кто придёт.

    Возвращает содержимое файла в виде байтов.
    """
    stmt = select(RSVP)
    if attending_only:
        stmt = stmt.where(RSVP.attending.is_(True))
    stmt = stmt.order_by(RSVP.full_name)
    records = list(db.scalars(stmt))

    tz = ZoneInfo(config.WEDDING_TZID)

    wb = Workbook()
    ws = wb.active
    ws.title = "Гости"

    # Заголовок.
    header_fill = PatternFill("solid", fgColor="A3162A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(_HEADER, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_people = 0
    for idx, rec in enumerate(records, start=1):
        people = 1 + (rec.guests_count or 0)
        total_people += people

        updated = rec.updated_at
        if updated is not None:
            if updated.tzinfo is None:
                updated = updated.replace(tzinfo=timezone.utc)
            updated_str = updated.astimezone(tz).strftime("%d.%m.%Y %H:%M")
        else:
            updated_str = ""

        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=rec.full_name)
        ws.cell(row=row, column=3, value=rec.guests_count)
        ws.cell(row=row, column=4, value=people)
        ws.cell(row=row, column=5, value=updated_str)

    # Итоговая строка.
    total_row = len(records) + 2
    label = ws.cell(row=total_row, column=2, value="Итого")
    label.font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=f"гостей: {len(records)}").font = Font(bold=True)
    people_cell = ws.cell(row=total_row, column=4, value=total_people)
    people_cell.font = Font(bold=True)

    # Ширины колонок.
    widths = [6, 34, 14, 16, 20]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
